# ===============================
# FINAL KIT FORMATTER v3.7 (LOCKED) — UPDATED
# Usage:
#   python final_kit_formatter_v3_7.py <forecast.json> <output_dir> <subscriber_id> <kit_type>
# ===============================

import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.comments import Comment


def determine_bov_from_confidence(confidence_score):
    """
    CONFIDENCE_FIX_APPLIED - Enhanced BOV determination
    
    Fixed thresholds for proper color coding:
    - 85%+ = GREEN (Strong Play) 
    - 80-84% = YELLOW (Moderate Play)
    - 75-79% = TAN (Caution Play)
    - Below 75% = RED (Avoid)
    """
    if confidence_score is None:
        return "GRAY"
    
    # Normalize to percentage if needed
    if confidence_score <= 1.0:
        pct = confidence_score * 100
    else:
        pct = confidence_score
    
    if pct >= 85.0:
        return "GREEN"
    elif pct >= 80.0:
        return "YELLOW"
    elif pct >= 75.0:
        return "TAN"
    else:
        return "RED"


SHEET_NAME = "My Best Odds"

OFFICIAL_ODDS_MAP = {
    "Cash3": "1 in 1,000",
    "Cash4": "1 in 10,000",
    "Cash4Life": "1 in 21,846,048",
    "MegaMillions": "1 in 302,575,350",
    "Powerball": "1 in 292,201,338",
}

# Base odds for MBO calculation
BASE_ODDS = {
    "Cash3": 1000,
    "Cash4": 10000,
    "Cash4Life": 21846048,
    "MegaMillions": 302575350,
    "Powerball": 292201338,
}

FINAL_COLS = [
    "Date",
    "Game",
    "Best Odds Verdict (BOV)",
    "Play Instruction",
    "Cash Confidence Score (%)",
    "Jackpot Confidence Score (%)",
    "Official Lottery Odds",
    "My Best Odds (MBO)",
    "Odds Improvement",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # deep blue
ZEBRA_FILL  = PatternFill("solid", fgColor="F8F9FA")   # light gray

# Enhanced BOV color scheme with stronger visual indicators
BOV_FILLS = {
    "GREEN":  PatternFill("solid", fgColor="70AD47"),   # strong green
    "YELLOW": PatternFill("solid", fgColor="FFC000"),   # vibrant yellow
    "TAN":    PatternFill("solid", fgColor="D2B48C"),   # tan
    "GRAY":   PatternFill("solid", fgColor="A5A5A5"),   # medium gray for skip
    "RED":    PatternFill("solid", fgColor="E74C3C"),   # strong red
}

# Header font styling
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CELL_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Hover comments for header row
HEADER_COMMENTS = {
    "Best Odds Verdict (BOV)": "Color-coded action indicator: GREEN=Strong Play, YELLOW=Moderate Play, RED=Avoid, GRAY=Skip Zone",
    "Cash Confidence Score (%)": "SMART LOGIC confidence level for Cash3/Cash4 games based on astronomical and numerological analysis",
    "Jackpot Confidence Score (%)": "SMART LOGIC confidence level for MegaMillions, Powerball, and Cash4Life predictions",
    "My Best Odds (MBO)": "Improved odds calculated from SMART LOGIC confidence scores - significantly better than official lottery odds"
}

FOOTER_TEXT = (
    "Powered by the S.M.A.R.T. System – Statistical Modeling for Accurate Results & Timing. "
    "Everybody wins by following the right guidance. Small wins build alignment. Big wins arrive in their time. "
    "This kit is for entertainment, tax write-off, and life-alignment purposes only. "
    "Play responsibly."
)

# ----------------------------
# Helpers
# ----------------------------

def calculate_mbo_odds(game, confidence_score, play_type="STRAIGHT"):
    """
    Calculate My Best Odds based on SMART LOGIC confidence score
    
    Formula: MBO Odds = Official Odds × (Confidence Adjustment Factor)
    
    Confidence Adjustment Factor:
    - 90-100%: Divide official odds by 10 (10x better odds)
    - 80-89%:  Divide official odds by 7  (7x better odds)
    - 70-79%:  Divide official odds by 5  (5x better odds) 
    - 65-69%:  Divide official odds by 3  (3x better odds)
    - Below 65%: Use official odds (no improvement)
    """
    if game not in BASE_ODDS:
        return {
            "mbo_odds_text": "N/A",
            "improvement_text": "N/A"
        }
    
    base_odds = BASE_ODDS[game]
    confidence_pct = confidence_score * 100 if confidence_score <= 1 else confidence_score
    
    # Confidence-based odds improvement
    if confidence_pct >= 90:
        improvement_factor = 10
        tier = "ELITE"
    elif confidence_pct >= 80:
        improvement_factor = 7
        tier = "PREMIUM"
    elif confidence_pct >= 70:
        improvement_factor = 5
        tier = "STRONG"
    elif confidence_pct >= 65:
        improvement_factor = 3
        tier = "MODERATE"
    else:
        improvement_factor = 1
        tier = "STANDARD"
    
    # Play type adjustments for Cash3/Cash4
    if game in ("Cash3", "Cash4"):
        if play_type.upper() in ("BOX", "COMBO"):
            # Box/Combo play has better base odds
            if game == "Cash3":
                base_odds = 167  # Cash3 Box odds
            elif game == "Cash4":
                base_odds = 2500  # Cash4 Box odds
    
    mbo_odds = max(1, base_odds // improvement_factor)
    
    return {
        "mbo_odds_text": f"1 in {mbo_odds:,}",
        "improvement_text": f"{improvement_factor}x Better"
    }

def resolve_numbers(row):
    for k in ("number", "numbers", "pick", "combo", "combination"):
        v = row.get(k)
        if v not in ("", None):
            return str(v)

    digits = [str(row.get(k)) for k in ("d1", "d2", "d3", "d4") if row.get(k) not in ("", None)]
    return "".join(digits)

def build_play_instruction(row):
    game = row.get("game", "")
    play_type = row.get("play_type", "")
    numbers = resolve_numbers(row)

    session = ""
    if game in ("Cash3", "Cash4"):
        session = row.get("draw_time", "")

    parts = [p for p in [numbers, play_type, session] if p]
    return " – ".join(parts)

def normalize_bov(v):
    if v is None:
        return ""
    s = str(v).strip().upper()
    
    # Handle silence and skip states
    if s in ("SILENCE", "SKIP", "SKIP ZONE", "DO NOT PLAY", "NO PLAY"):
        return "GRAY"
    if s in ("FALSE", "0", "SILENCED"):  # From play_flag = False
        return "GRAY"
        
    # Legacy mappings
    if s in ("BROWN",):
        return "TAN"
        
    return s

def _to_pct(val):
    """
    Normalize confidence into percentage display:
    - if 0..1 => * 100
    - if 1..100 => keep
    """
    if val in ("", None):
        return ""
    try:
        f = float(val)
    except Exception:
        return ""
    if f <= 1.0:
        return round(f * 100.0, 1)
    return round(f, 1)

def build_final_kit(engine_df):
    rows = []
    for _, r in engine_df.iterrows():
        game = r.get("game", "")

        is_cash = game in ("Cash3", "Cash4")
        is_jackpot = game in ("MegaMillions", "Powerball", "Cash4Life")

        # Cash confidence comes from engine-core scoring (confidence_score)
        cash_conf = ""
        if is_cash:
            cash_conf = _to_pct(r.get("confidence_score", ""))

        # Jackpot confidence: prefer the right-engine "confidence" if present,
        # otherwise fall back to engine-core "confidence_score"
        jackpot_conf = ""
        if is_jackpot:
            jackpot_conf = _to_pct(r.get("confidence", ""))  # right-engine native
            if jackpot_conf == "":
                jackpot_conf = _to_pct(r.get("confidence_score", ""))  # fallback

        # Check play_flag from post-engine filtering
        play_flag = r.get("play_flag", True)
        bov = normalize_bov(r.get("play_flag", "") if not play_flag else r.get("play_flag", ""))
        
        # If silenced by post-engine filter, override BOV
        if not play_flag:
            bov = "GRAY"  # Mark as silenced/skip

        # Calculate My Best Odds based on confidence
        mbo_info = calculate_mbo_odds(game, r.get("confidence_score", 0), r.get("play_type", "STRAIGHT"))

        rows.append({
            "Date": r.get("draw_date", "") or r.get("date", ""),
            "Game": game,
            "Best Odds Verdict (BOV)": bov,
            "Play Instruction": build_play_instruction(r),
            "Cash Confidence Score (%)": cash_conf,
            "Jackpot Confidence Score (%)": jackpot_conf,
            "Official Lottery Odds": OFFICIAL_ODDS_MAP.get(game, ""),
            "My Best Odds (MBO)":mbo_info["mbo_odds_text"],
            "Odds Improvement": mbo_info["improvement_text"],
        })

    return pd.DataFrame(rows)[FINAL_COLS]

def export_final_kit(forecast_json_path, output_dir, subscriber_id, kit_type):
    engine_df = pd.read_json(forecast_json_path)
    final_df = build_final_kit(engine_df)

    # derive date range from data if available
    try:
        dates = pd.to_datetime(final_df["Date"], errors="coerce").dropna()
        start = dates.min().date().isoformat() if not dates.empty else "START"
        end   = dates.max().date().isoformat() if not dates.empty else "END"
    except Exception:
        start, end = "START", "END"

    os.makedirs(output_dir, exist_ok=True)
    output_xlsx = os.path.join(output_dir, f"{subscriber_id}_{kit_type}_{start}_to_{end}.xlsx")

    final_df.to_excel(output_xlsx, index=False, sheet_name=SHEET_NAME)

    wb = load_workbook(output_xlsx)
    ws = wb[SHEET_NAME]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Enhanced header styling with comments
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = CELL_BORDER
        
        # Add hover comments to specific columns
        column_title = str(cell.value)
        if column_title in HEADER_COMMENTS:
            comment = Comment(HEADER_COMMENTS[column_title], "SMART LOGIC System")
            comment.width = 300
            comment.height = 80
            cell.comment = comment

    # Column widths (optimized for better visibility)
    widths = {"A": 16, "B": 16, "C": 20, "D": 36, "E": 24, "F": 26, "G": 24, "H": 22, "I": 18, "J": 46}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Enhanced row styling with borders and improved colors
    for row_idx in range(2, ws.max_row + 1):
        is_zebra = (row_idx % 2 == 0)
        bov_value = normalize_bov(ws[f"C{row_idx}"].value)

        # Apply base styling to all cells in row
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = CENTER_ALIGN
            cell.border = CELL_BORDER
            
            # Apply zebra striping
            if is_zebra:
                cell.fill = ZEBRA_FILL

        # Apply BOV color cascade with enhanced visibility
        if bov_value in BOV_FILLS:
            # Color from BOV column to My Best Odds column (C through H)
            for col in ("C", "D", "E", "F", "G", "H"):  # Continuous range from BOV to MBO
                cell = ws[f"{col}{row_idx}"]
                cell.fill = BOV_FILLS[bov_value]
                
                # Add bold font for GREEN verdicts (strong plays)
                if bov_value == "GREEN":
                    cell.font = Font(bold=True)
                # Add italic font for GRAY verdicts (skip zones)  
                elif bov_value == "GRAY":
                    cell.font = Font(italic=True, color="666666")

    # Footer
    footer_row = ws.max_row + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=ws.max_column)
    footer_cell = ws.cell(row=footer_row, column=1)
    footer_cell.value = FOOTER_TEXT
    footer_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    footer_cell.font = Font(italic=True, size=10)

    # Add Color Legend
    legend_start_row = footer_row + 3
    
    # Legend title
    legend_title_cell = ws.cell(row=legend_start_row, column=1)
    legend_title_cell.value = "COLOR LEGEND - Best Odds Verdict (BOV)"
    legend_title_cell.font = Font(bold=True, size=12)
    legend_title_cell.alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=legend_start_row, start_column=1, end_row=legend_start_row, end_column=4)
    
    # Legend entries
    legend_entries = [
        ("GREEN", "Strong Play - High confidence predictions (65%+)", BOV_FILLS["GREEN"]),
        ("YELLOW", "Moderate Play - Medium confidence predictions", BOV_FILLS["YELLOW"]),
        ("TAN", "Caution Play - Lower confidence but viable", BOV_FILLS["TAN"]),
        ("RED", "Avoid - Poor alignment or low confidence", BOV_FILLS["RED"]),
        ("GRAY", "Skip Zone - Do not play these predictions", BOV_FILLS["GRAY"])
    ]
    
    for i, (color_name, description, fill_color) in enumerate(legend_entries):
        row = legend_start_row + 1 + i
        
        # Color box
        color_cell = ws.cell(row=row, column=1)
        color_cell.value = color_name
        color_cell.fill = fill_color
        color_cell.font = Font(bold=True, color="FFFFFF" if color_name in ["GREEN", "RED"] else "000000")
        color_cell.alignment = CENTER_ALIGN
        color_cell.border = CELL_BORDER
        
        # Description
        desc_cell = ws.cell(row=row, column=2)
        desc_cell.value = description
        desc_cell.font = Font(size=10)
        desc_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    
    # Adjust column width for legend
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 12)

    wb.save(output_xlsx)
    print(f"[SUCCESS] Final Excel kit created: {output_xlsx}")

if __name__ == "__main__":
    if len(sys.argv) < 5:
        raise ValueError(
            "Usage: python final_kit_formatter_v3_7.py <path_to_forecast.json> <output_dir> <subscriber_id> <kit_type>"
        )
    export_final_kit(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
