import sys
import json
import csv
import re
from pathlib import Path
from collections import defaultdict
from statistics import mean
from datetime import datetime, timedelta
from typing import Optional

from openpyxl import load_workbook

# ==============================
# CONFIG – TWEAK AS NEEDED
# ==============================

PROJECT_ROOT = Path(__file__).resolve().parent
OUTPUT_ROOT = PROJECT_ROOT / "outputs"
GA_RESULTS_DIR = PROJECT_ROOT / "data" / "ga_results"

# Date window for this audit
COVERAGE_START = "2025-09-01"
COVERAGE_END = "2025-11-10"

# Mapping between internal game names and keys in summary.json
GAME_KEYS = {
    "Cash3": "Cash3",
    "Cash4": "Cash4",
    "MegaMillions": "MegaMillions",
    "Powerball": "Powerball",
    "Cash4Life": "Cash4Life",
}

DATE_FMT = "%Y-%m-%d"


# ==============================
# UTILITIES – FIXED & UPGRADED
# ==============================

def smart_parse_date(s):
    """Parse ANY date format, including Excel serials."""
    if s is None:
        return None

    s = str(s).strip()

    # Excel serial numbers (e.g., 45000 → date)
    if s.isdigit() and int(s) > 30000:
        return (datetime(1899, 12, 30) + timedelta(days=int(s))).date()

    # Already ISO
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        pass

    # Common messy formats
    fmts = [
        "%m/%d/%Y",
        "%m/%d/%y",
        "%m-%d-%Y",
        "%m-%d-%y",
        "%Y/%m/%d",
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%b %d %Y",
        "%B %d %Y",
    ]

    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue

    # Final attempt
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


def normalize_date_string(s):
    d = smart_parse_date(s)
    return d.strftime("%Y-%m-%d") if d else None


def parse_date(s):
    return smart_parse_date(s)


def within_window(dstr):
    nd = normalize_date_string(dstr)
    if not nd:
        return False
    d = parse_date(nd)
    return parse_date(COVERAGE_START) <= d <= parse_date(COVERAGE_END)


def load_json(path: Path):
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_csv(path: Path):
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def load_excel_as_dicts(path: Path, sheet_name: Optional[str] = None):
    """Load an Excel worksheet as list of dict rows."""
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]

    out = []
    for r in rows[1:]:
        row_dict = {}
        for h, v in zip(headers, r):
            if h:
                row_dict[h] = v
        out.append(row_dict)

    return out


# ==============================
# LOAD OFFICIAL RESULTS – CASH3 / CASH4
# ==============================

def _parse_cash_number_with_leading_zeros(val, num_digits: int):
    """
    Ensure we keep leading zeros for Cash3/Cash4.
    """
    if val is None:
        return None
    s = str(val).strip().replace(" ", "")
    if not s:
        return None
    s = s.zfill(num_digits)
    if len(s) != num_digits:
        return None
    return tuple(int(ch) for ch in s)


def load_cash3_midnight_from_template():
    """
    Cash3 Midday & Night from mGA_Cash3_Cash4_Manual_Entry_Template.xlsx
    """
    results = {}
    template_xlsx = GA_RESULTS_DIR / "mGA_Cash3_Cash4_Manual_Entry_Template.xlsx"
    if not template_xlsx.exists():
        print(f"[WARNING] Cash3/Cash4 template Excel not found at {template_xlsx}")
        return results

    rows = load_excel_as_dicts(template_xlsx)
    for row in rows:
        game = str(row.get("Game") or "").strip().lower()
        if game not in ("cash 3", "cash3"):
            continue

        date_val = row.get("Draw Date")
        if not date_val:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.date().isoformat()
        else:
            date_str = str(date_val)[:10]

        if not within_window(date_str):
            continue

        wn = row.get("Winning Numbers")
        digits = _parse_cash_number_with_leading_zeros(wn, 3)
        if digits is None:
            print(f"[WARNING] Skipping Cash3 Mid/Night row with invalid digits '{wn}' at {date_str}")
            continue

        if date_str not in results:
            results[date_str] = {"main": digits, "bonus": None}

    print(f"[INFO] Loaded {len(results)} Cash3 Midday/Night dates from template.")
    return results


def load_cash4_midnight_from_template():
    """
    Cash4 Midday & Night from mGA_Cash3_Cash4_Manual_Entry_Template.xlsx
    """
    results = {}
    template_xlsx = GA_RESULTS_DIR / "mGA_Cash3_Cash4_Manual_Entry_Template.xlsx"
    if not template_xlsx.exists():
        print(f"[WARNING] Cash3/Cash4 template Excel not found at {template_xlsx}")
        return results

    rows = load_excel_as_dicts(template_xlsx)
    for row in rows:
        game = str(row.get("Game") or "").strip().lower()
        if game not in ("cash 4", "cash4"):
            continue

        date_val = row.get("Draw Date")
        if not date_val:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.date().isoformat()
        else:
            date_str = str(date_val)[:10]

        if not within_window(date_str):
            continue

        wn = row.get("Winning Numbers")
        digits = _parse_cash_number_with_leading_zeros(wn, 4)
        if digits is None:
            print(f"[WARNING] Skipping Cash4 Mid/Night row with invalid digits '{wn}' at {date_str}")
            continue

        if date_str not in results:
            results[date_str] = {"main": digits, "bonus": None}

    print(f"[INFO] Loaded {len(results)} Cash4 Midday/Night dates from template.")
    return results


def load_cash3_evening_csv():
    """
    Cash3 Evening from 'Cash3 Evening 901-1110 (1).csv'.
    """
    results = {}
    path = GA_RESULTS_DIR / "Cash3 Evening 901-1110 (1).csv"
    if not path.exists():
        print(f"[WARNING] Cash3 Evening CSV not found at {path}")
        return results

    rows = load_csv(path)
    for row in rows:
        date_val = row.get("Draw Date") or row.get("Date") or row.get("date")
        if not date_val:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.date().isoformat()
        else:
            date_str = str(date_val)[:10]

        if not within_window(date_str):
            continue

        wn = (
            row.get("Winning Numbers")
            or row.get("Digits")
            or row.get("Number")
            or row.get("Result")
        )
        if wn is None:
            continue

        digits = _parse_cash_number_with_leading_zeros(wn, 3)
        if digits is None:
            print(f"[WARNING] Skipping Cash3 Evening row with invalid digits '{wn}' at {date_str}")
            continue

        if date_str not in results:
            results[date_str] = {"main": digits, "bonus": None}

    print(f"[INFO] Loaded {len(results)} Cash3 Evening dates from CSV.")
    return results


def load_cash4_evening_results():
    """
    Cash4 EVENING draws from 'Cash 4 Evening.xlsx'.
    """
    results = {}
    path = GA_RESULTS_DIR / "Cash 4 Evening.xlsx"
    if not path.exists():
        print(f"[WARNING] Cash4 Evening file not found at {path}")
        return results

    rows = load_excel_as_dicts(path)
    for row in rows:
        date_val = row.get("Cash 4")
        if not date_val:
            continue

        if isinstance(date_val, datetime):
            date_str = date_val.date().isoformat()
        else:
            date_str = str(date_val)[:10]

        if not within_window(date_str):
            continue

        time_val = row.get("Time")
        if not time_val:
            continue

        time_str = str(time_val)
        cleaned = "".join(time_str.split())
        if len(cleaned) != 4:
            print(f"[WARNING] Unexpected Cash4 Evening digits '{time_val}' at {date_str}")
            continue

        digits = tuple(int(ch) for ch in cleaned)
        if date_str not in results:
            results[date_str] = {"main": digits, "bonus": None}

    print(f"[INFO] Loaded {len(results)} Cash4 Evening dates from Excel.")
    return results


def load_cash3_results():
    combined = {}
    mid = load_cash3_midnight_from_template()
    eve = load_cash3_evening_csv()

    for d, v in mid.items():
        combined.setdefault(d, v)
    for d, v in eve.items():
        combined.setdefault(d, v)

    print(f"[INFO] Combined Cash3 dates: {len(combined)}")
    return combined


def load_cash4_results():
    combined = {}
    mid = load_cash4_midnight_from_template()
    eve = load_cash4_evening_results()

    for d, v in mid.items():
        combined.setdefault(d, v)
    for d, v in eve.items():
        combined.setdefault(d, v)

    print(f"[INFO] Combined Cash4 dates: {len(combined)}")
    return combined


# ==============================
# LOAD OFFICIAL RESULTS – JACKPOT GAMES
# ==============================

def extract_main_and_bonus_from_string(s, expected_main_count: int = 5):
    """
    Parse messy jackpot strings like:
      '28 32 36 51 69 Powerball:02 Power Play:2X'
      '09 21 27 48 56'
      '10,24,27,42,51'
    Returns (main_tuple, bonus_int_or_None).
    """
    if s is None:
        return None, None

    text = str(s)
    text = text.replace(",", " ")
    nums = re.findall(r"\d+", text)
    if not nums:
        return None, None

    ints = [int(n) for n in nums]

    if len(ints) < expected_main_count:
        return None, None

    main = tuple(ints[:expected_main_count])
    bonus = ints[expected_main_count] if len(ints) > expected_main_count else None
    return main, bonus


def load_jackpot_from_xlsx(
    filename: str,
    game_name: str,
    bonus_field_candidates=None,
    sheet_name: Optional[str] = None,
):
    if bonus_field_candidates is None:
        bonus_field_candidates = [
            "MegaBall",
            "Megaball",
            "MB",
            "Powerball",
            "PB",
            "CashBall",
            "Cash Ball",
            "CB",
        ]

    path = GA_RESULTS_DIR / filename
    results = {}

    if not path.exists():
        print(f"[WARNING] Jackpot Excel not found for {game_name} at {path}")
        return results

    rows = load_excel_as_dicts(path, sheet_name=sheet_name)

    for row in rows:
        date_val = (
            row.get("Date")
            or row.get("DATE")
            or row.get("Draw Date")
            or row.get("Draw date")
            or row.get("DrawDate")
        )
        if not date_val:
            continue

        date_str = normalize_date_string(date_val)
        if not date_str:
            continue
        if not within_window(date_str):
            continue

        wn = (
            row.get("Winning Numbers")
            or row.get("WinningNumbers")
            or row.get("Numbers")
            or row.get("Result")
        )
        if wn is None:
            continue

        bonus = None
        for bname in bonus_field_candidates:
            if bname in row and row[bname] not in (None, ""):
                raw_bonus = row[bname]
                nums = re.findall(r"\d+", str(raw_bonus))
                if nums:
                    bonus = int(nums[0])
                break

        main_from_str, bonus_from_str = extract_main_and_bonus_from_string(
            wn, expected_main_count=5
        )

        if main_from_str is None:
            print(
                f"[WARNING] Could not parse main numbers for {game_name} on {date_str} from '{wn}'"
            )
            continue

        if bonus is None:
            bonus = bonus_from_str

        results[date_str] = {"main": main_from_str, "bonus": bonus}

    print(f"[INFO] Loaded {len(results)} {game_name} draws from {filename}")
    return results


def load_all_official_results():
    official = {}

    official["Cash3"] = load_cash3_results()
    official["Cash4"] = load_cash4_results()

    official["MegaMillions"] = load_jackpot_from_xlsx(
        "Mega Millions 9-10-25 - 11-10-25.xlsx",
        game_name="MegaMillions",
        bonus_field_candidates=["Cash Ball", "CashBall", "MegaBall", "Megaball", "MB"],
    )

    official["Powerball"] = load_jackpot_from_xlsx(
        "Powerball.xlsx",
        game_name="Powerball",
        bonus_field_candidates=["Powerball", "PB"],
    )

    official["Cash4Life"] = load_jackpot_from_xlsx(
        "Cash4Life 9-01-2025-11-10-25.xlsx",
        game_name="Cash4Life",
        bonus_field_candidates=["Cash Ball", "CashBall", "CB"],
    )

    return official


# ==============================
# HIT CHECKING
# ==============================

def normalize_int_list(nums):
    """Convert list of str/int to list of ints."""
    out = []
    for n in nums:
        out.append(int(n))
    return out


def check_cash_hit(pred, actual):
    """
    For Cash3/Cash4:
    Returns (straight_hit: bool, box_hit: bool)
    """
    if pred is None or actual is None:
        return False, False
    pred_t = tuple(normalize_int_list(pred))
    act_t = tuple(actual)
    straight = (pred_t == act_t)
    box = sorted(pred_t) == sorted(act_t)
    return straight, box


def check_jackpot_hit(pred_main, pred_bonus, act_main, act_bonus):
    """
    For MM, PB, C4L:
    - Jackpot full match = all main numbers match (order-insensitive) AND bonus match.
    - Returns (jackpot_full, main_match_count, bonus_match)
    """
    if pred_main is None or act_main is None:
        return False, 0, False

    pset = set(normalize_int_list(pred_main))
    aset = set(act_main)
    main_match_count = len(pset & aset)
    bonus_match = False

    if pred_bonus is not None and act_bonus is not None:
        bonus_match = int(pred_bonus) == int(act_bonus)

    jackpot_full = (main_match_count == len(aset)) and bonus_match
    return jackpot_full, main_match_count, bonus_match


# ==============================
# MAIN AUDIT LOGIC (GENERALIZED)
# ==============================

def find_kit_summaries(kit_name: str, initials: str):
    """
    Find summary.json files for a specific kit + initials in the coverage window.
    Pattern: <KIT>_<INITIALS>_2025-09-01_to_2025-11-10/summary.json
    """
    pattern = f"{kit_name}_{initials}_{COVERAGE_START}_to_{COVERAGE_END}"
    paths = []
    for folder in OUTPUT_ROOT.glob(pattern):
        summary_path = folder / "summary.json"
        if summary_path.exists():
            paths.append(summary_path)
    return paths


def audit_kit(kit_name: str, initials: str):
    official = load_all_official_results()
    summary_paths = find_kit_summaries(kit_name, initials)
    print(f"[INFO] Found {len(summary_paths)} {kit_name} summary files for {initials} to audit.")

    audit_rows = []
    jackpot_hits = []
    kit_performance = defaultdict(lambda: {"total_hits": 0, "jackpot_hits": 0, "days": 0})
    subscriber_perf = {}

    for summary_path in summary_paths:
        summary = load_json(summary_path)
        subscriber_info = summary.get("subscriber", {})
        # If name/initials missing in JSON, infer from folder
        folder_name = summary_path.parent.name
        inferred_initials = folder_name.split("_")[1] if "_" in folder_name else initials

        sub_initials = subscriber_info.get("initials") or inferred_initials
        name = subscriber_info.get("name", sub_initials)

        sub_key = f"{kit_name}:{sub_initials}"
        sub_stats = subscriber_perf.setdefault(sub_key, {
            "subscriber": name,
            "initials": sub_initials,
            "kit": kit_name,
            "total_hits": 0,
            "cash3_hits": 0,
            "cash4_hits": 0,
            "mm_hits": 0,
            "pb_hits": 0,
            "c4l_hits": 0,
            "jackpot_hits": 0,
            "dates_with_hits": set(),
            "scores_on_hits": [],
        })

        days = summary.get("days", [])
        for day in days:
            date_str = day.get("date")
            if not date_str or not within_window(date_str):
                continue

            scores = day.get("scores", {})
            total_score = scores.get("total", None)
            picks = day.get("picks", {})

            # Count days for kit stats (unique dates)
            kit_performance[kit_name]["days"] += 1

            # ---- Cash3 ----
            c3_key = GAME_KEYS["Cash3"]
            if c3_key in picks and date_str in official["Cash3"]:
                pred_c3 = picks[c3_key]
                if isinstance(pred_c3, dict):
                    pred_c3 = pred_c3.get("digits") or pred_c3.get("nums")
                act_c3 = official["Cash3"][date_str]["main"]
                c3_straight, c3_box = check_cash_hit(pred_c3, act_c3)

                audit_rows.append({
                    "kit": kit_name,
                    "subscriber": name,
                    "initials": sub_initials,
                    "date": date_str,
                    "game": "Cash3",
                    "pred": " ".join(str(x) for x in normalize_int_list(pred_c3)) if pred_c3 else "",
                    "actual": " ".join(str(x) for x in act_c3),
                    "straight_hit": int(c3_straight),
                    "box_hit": int(c3_box),
                    "jackpot_full": 0,
                    "main_match_count": "",
                    "bonus_match": "",
                    "score": total_score if total_score is not None else "",
                })

                if c3_straight or c3_box:
                    sub_stats["total_hits"] += 1
                    sub_stats["cash3_hits"] += 1
                    sub_stats["dates_with_hits"].add(date_str)
                    if total_score is not None:
                        sub_stats["scores_on_hits"].append(float(total_score))
                    kit_performance[kit_name]["total_hits"] += 1

            # ---- Cash4 ----
            c4_key = GAME_KEYS["Cash4"]
            if c4_key in picks and date_str in official["Cash4"]:
                pred_c4 = picks[c4_key]
                if isinstance(pred_c4, dict):
                    pred_c4 = pred_c4.get("digits") or pred_c4.get("nums")
                act_c4 = official["Cash4"][date_str]["main"]
                c4_straight, c4_box = check_cash_hit(pred_c4, act_c4)

                audit_rows.append({
                    "kit": kit_name,
                    "subscriber": name,
                    "initials": sub_initials,
                    "date": date_str,
                    "game": "Cash4",
                    "pred": " ".join(str(x) for x in normalize_int_list(pred_c4)) if pred_c4 else "",
                    "actual": " ".join(str(x) for x in act_c4),
                    "straight_hit": int(c4_straight),
                    "box_hit": int(c4_box),
                    "jackpot_full": 0,
                    "main_match_count": "",
                    "bonus_match": "",
                    "score": total_score if total_score is not None else "",
                })

                if c4_straight or c4_box:
                    sub_stats["total_hits"] += 1
                    sub_stats["cash4_hits"] += 1
                    sub_stats["dates_with_hits"].add(date_str)
                    if total_score is not None:
                        sub_stats["scores_on_hits"].append(float(total_score))
                    kit_performance[kit_name]["total_hits"] += 1

            # ---- MegaMillions ----
            mm_key = GAME_KEYS["MegaMillions"]
            if mm_key in picks and date_str in official["MegaMillions"]:
                pred_mm = picks[mm_key]
                if isinstance(pred_mm, dict):
                    pred_main = pred_mm.get("main") or pred_mm.get("numbers")
                    pred_bonus = pred_mm.get("bonus") or pred_mm.get("megaball")
                else:
                    pred_main = pred_mm[:5]
                    pred_bonus = pred_mm[5] if len(pred_mm) > 5 else None

                act_mm_main = official["MegaMillions"][date_str]["main"]
                act_mm_bonus = official["MegaMillions"][date_str]["bonus"]

                mm_jackpot, mm_main_match, mm_bonus_match = check_jackpot_hit(
                    pred_main, pred_bonus, act_mm_main, act_mm_bonus
                )

                audit_rows.append({
                    "kit": kit_name,
                    "subscriber": name,
                    "initials": sub_initials,
                    "date": date_str,
                    "game": "MegaMillions",
                    "pred": " ".join(str(x) for x in normalize_int_list(pred_main)) if pred_main else "",
                    "actual": " ".join(str(x) for x in act_mm_main) + (
                        f" + {act_mm_bonus}" if act_mm_bonus is not None else ""
                    ),
                    "straight_hit": "",
                    "box_hit": "",
                    "jackpot_full": int(mm_jackpot),
                    "main_match_count": mm_main_match,
                    "bonus_match": int(mm_bonus_match),
                    "score": total_score if total_score is not None else "",
                })

                if mm_jackpot:
                    sub_stats["total_hits"] += 1
                    sub_stats["mm_hits"] += 1
                    sub_stats["jackpot_hits"] += 1
                    kit_performance[kit_name]["jackpot_hits"] += 1
                    kit_performance[kit_name]["total_hits"] += 1
                    sub_stats["dates_with_hits"].add(date_str)
                    if total_score is not None:
                        sub_stats["scores_on_hits"].append(float(total_score))

                    jackpot_hits.append({
                        "kit": kit_name,
                        "subscriber": name,
                        "initials": sub_initials,
                        "date": date_str,
                        "game": "MegaMillions",
                        "pred_main": normalize_int_list(pred_main) if pred_main else [],
                        "pred_bonus": pred_bonus,
                        "actual_main": list(act_mm_main),
                        "actual_bonus": act_mm_bonus,
                        "score": total_score,
                    })

            # ---- Powerball ----
            pb_key = GAME_KEYS["Powerball"]
            if pb_key in picks and date_str in official["Powerball"]:
                pred_pb = picks[pb_key]
                if isinstance(pred_pb, dict):
                    pred_main = pred_pb.get("main") or pred_pb.get("numbers")
                    pred_bonus = pred_pb.get("bonus") or pred_pb.get("powerball")
                else:
                    pred_main = pred_pb[:5]
                    pred_bonus = pred_pb[5] if len(pred_pb) > 5 else None

                act_pb_main = official["Powerball"][date_str]["main"]
                act_pb_bonus = official["Powerball"][date_str]["bonus"]

                pb_jackpot, pb_main_match, pb_bonus_match = check_jackpot_hit(
                    pred_main, pred_bonus, act_pb_main, act_pb_bonus
                )

                audit_rows.append({
                    "kit": kit_name,
                    "subscriber": name,
                    "initials": sub_initials,
                    "date": date_str,
                    "game": "Powerball",
                    "pred": " ".join(str(x) for x in normalize_int_list(pred_main)) if pred_main else "",
                    "actual": " ".join(str(x) for x in act_pb_main) + (
                        f" + {act_pb_bonus}" if act_pb_bonus is not None else ""
                    ),
                    "straight_hit": "",
                    "box_hit": "",
                    "jackpot_full": int(pb_jackpot),
                    "main_match_count": pb_main_match,
                    "bonus_match": int(pb_bonus_match),
                    "score": total_score if total_score is not None else "",
                })

                if pb_jackpot:
                    sub_stats["total_hits"] += 1
                    sub_stats["pb_hits"] += 1
                    sub_stats["jackpot_hits"] += 1
                    kit_performance[kit_name]["jackpot_hits"] += 1
                    kit_performance[kit_name]["total_hits"] += 1
                    sub_stats["dates_with_hits"].add(date_str)
                    if total_score is not None:
                        sub_stats["scores_on_hits"].append(float(total_score))

                    jackpot_hits.append({
                        "kit": kit_name,
                        "subscriber": name,
                        "initials": sub_initials,
                        "date": date_str,
                        "game": "Powerball",
                        "pred_main": normalize_int_list(pred_main) if pred_main else [],
                        "pred_bonus": pred_bonus,
                        "actual_main": list(act_pb_main),
                        "actual_bonus": act_pb_bonus,
                        "score": total_score,
                    })

            # ---- Cash4Life ----
            c4l_key = GAME_KEYS["Cash4Life"]
            if c4l_key in picks and date_str in official["Cash4Life"]:
                pred_c4l = picks[c4l_key]
                if isinstance(pred_c4l, dict):
                    pred_main = pred_c4l.get("main") or pred_c4l.get("numbers")
                    pred_bonus = pred_c4l.get("bonus") or pred_c4l.get("cash_ball")
                else:
                    pred_main = pred_c4l[:5]
                    pred_bonus = pred_c4l[5] if len(pred_c4l) > 5 else None

                act_c4l_main = official["Cash4Life"][date_str]["main"]
                act_c4l_bonus = official["Cash4Life"][date_str]["bonus"]

                c4l_jackpot, c4l_main_match, c4l_bonus_match = check_jackpot_hit(
                    pred_main, pred_bonus, act_c4l_main, act_c4l_bonus
                )

                audit_rows.append({
                    "kit": kit_name,
                    "subscriber": name,
                    "initials": sub_initials,
                    "date": date_str,
                    "game": "Cash4Life",
                    "pred": " ".join(str(x) for x in normalize_int_list(pred_main)) if pred_main else "",
                    "actual": " ".join(str(x) for x in act_c4l_main) + (
                        f" + {act_c4l_bonus}" if act_c4l_bonus is not None else ""
                    ),
                    "straight_hit": "",
                    "box_hit": "",
                    "jackpot_full": int(c4l_jackpot),
                    "main_match_count": c4l_main_match,
                    "bonus_match": int(c4l_bonus_match),
                    "score": total_score if total_score is not None else "",
                })

                if c4l_jackpot:
                    sub_stats["total_hits"] += 1
                    sub_stats["c4l_hits"] += 1
                    sub_stats["jackpot_hits"] += 1
                    kit_performance[kit_name]["jackpot_hits"] += 1
                    kit_performance[kit_name]["total_hits"] += 1
                    sub_stats["dates_with_hits"].add(date_str)
                    if total_score is not None:
                        sub_stats["scores_on_hits"].append(float(total_score))

                    jackpot_hits.append({
                        "kit": kit_name,
                        "subscriber": name,
                        "initials": sub_initials,
                        "date": date_str,
                        "game": "Cash4Life",
                        "pred_main": normalize_int_list(pred_main) if pred_main else [],
                        "pred_bonus": pred_bonus,
                        "actual_main": list(act_c4l_main),
                        "actual_bonus": act_c4l_bonus,
                        "score": total_score,
                    })

    # Convert date sets to counts and score averages
    for sub_key, stats in subscriber_perf.items():
        stats["unique_hit_days"] = len(stats["dates_with_hits"])
        if stats["scores_on_hits"]:
            stats["avg_score_on_hits"] = round(mean(stats["scores_on_hits"]), 2)
        else:
            stats["avg_score_on_hits"] = None
        stats["dates_with_hits"] = sorted(list(stats["dates_with_hits"]))

    return audit_rows, jackpot_hits, kit_performance, subscriber_perf


# ==============================
# EXPORT HELPERS
# ==============================

def write_csv(path: Path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


def main():
    if len(sys.argv) < 3:
        print("Usage: python audit_kit_v3.py <INITIALS> <KITNAME>")
        print("Example: python audit_kit_v3.py JDS BOOK3")
        sys.exit(1)

    initials = sys.argv[1].strip()
    kit_name = sys.argv[2].strip().upper()  # BOOK, BOOK3, BOSK, etc.

    audit_rows, jackpot_hits, kit_performance, subscriber_perf = audit_kit(kit_name, initials)

    audit_dir = OUTPUT_ROOT / f"AUDIT_{kit_name}_{initials}_{COVERAGE_START}_to_{COVERAGE_END}"
    audit_dir.mkdir(parents=True, exist_ok=True)

    # 1) Full audit CSV
    audit_csv_path = audit_dir / "audit_results.csv"
    fieldnames = [
        "kit",
        "subscriber",
        "initials",
        "date",
        "game",
        "pred",
        "actual",
        "straight_hit",
        "box_hit",
        "jackpot_full",
        "main_match_count",
        "bonus_match",
        "score",
    ]
    write_csv(audit_csv_path, audit_rows, fieldnames)
    print(f"[INFO] Wrote full audit CSV to {audit_csv_path}")

    # 2) Jackpot hits log + JSON
    jackpot_log_path = audit_dir / "jackpot_full_hits.log"
    with jackpot_log_path.open("w", encoding="utf-8") as f:
        for hit in jackpot_hits:
            line = (
                f"{hit['date']} | {hit['game']} | {hit['kit']} | "
                f"{hit['subscriber']} ({hit['initials']}) | "
                f"Pred: {hit['pred_main']} + {hit['pred_bonus']} | "
                f"Actual: {hit['actual_main']} + {hit['actual_bonus']} | "
                f"Score: {hit['score']}"
            )
            f.write(line + "\n")
    print(f"[INFO] Wrote jackpot hits log to {jackpot_log_path}")

    jackpot_json_path = audit_dir / "jackpot_full_hits.json"
    with jackpot_json_path.open("w", encoding="utf-8") as f:
        json.dump(jackpot_hits, f, indent=2)
    print(f"[INFO] Wrote jackpot hits JSON to {jackpot_json_path}")

    # 3) Kit performance summary
    kit_perf_json_path = audit_dir / "kit_performance_summary.json"
    with kit_perf_json_path.open("w", encoding="utf-8") as f:
        json.dump(kit_performance, f, indent=2)
    print(f"[INFO] Wrote kit performance summary to {kit_perf_json_path}")

    # 4) Subscriber performance summary
    sub_perf_json_path = audit_dir / "subscriber_performance.json"
    with sub_perf_json_path.open("w", encoding="utf-8") as f:
        json.dump(subscriber_perf, f, indent=2)
    print(f"[INFO] Wrote subscriber performance summary to {sub_perf_json_path}")


if __name__ == "__main__":
    main()
